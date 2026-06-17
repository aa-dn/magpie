#!/usr/bin/env python3
"""
Reverse Image Search Tool
Queries SerpAPI's Google Reverse Image Search and exports results to CSV, Excel (with
embedded thumbnails), and HTML (with inline images).

Usage:
    python reverse_image_search.py <image_url> [--api-key KEY] [--output PREFIX]
    python reverse_image_search.py <image_url> --local  # if image_url is a local file path
"""

import os
import sys
import csv
import json
import argparse
import requests
from io import BytesIO


# ── SerpAPI ───────────────────────────────────────────────────────────────────

def _call_engine(params: dict) -> dict:
    try:
        resp = requests.get("https://serpapi.com/search", params=params, timeout=30)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        return {"_request_error": str(e)}


def _parse_engine_results(data: dict, engine_label: str) -> list[dict]:
    items = (
        data.get("visual_matches") or
        data.get("image_results") or
        data.get("organic_results") or
        []
    )
    results = []
    for item in items:
        thumb = item.get("thumbnail", "")
        if isinstance(thumb, dict):
            thumb = thumb.get("src", "")
        results.append({
            "title":     item.get("title", ""),
            "url":       item.get("link") or item.get("url", ""),
            "source":    item.get("source", ""),
            "thumbnail": thumb or "",
            "engine":    engine_label,
        })
    return results


def search_reverse_image(image_url: str, api_key: str) -> dict:
    return _call_engine({"engine": "google_lens", "url": image_url, "api_key": api_key})


def search_all_engines(image_url: str, api_key: str) -> list[dict]:
    from concurrent.futures import ThreadPoolExecutor, as_completed

    engines = [
        ({"engine": "google_lens", "url": image_url, "api_key": api_key}, "Google Lens"),
        ({"engine": "yandex_reverse_image_search", "image_url": image_url, "api_key": api_key}, "Yandex"),
        ({"engine": "bing_visual_search", "image_url": image_url, "api_key": api_key}, "Bing"),
    ]

    by_url = {}

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {executor.submit(_call_engine, params): label for params, label in engines}
        for future in as_completed(futures):
            label = futures[future]
            data = future.result()
            if "_request_error" not in data:
                for r in _parse_engine_results(data, label):
                    if not r["url"]:
                        continue
                    if r["url"] in by_url:
                        existing = by_url[r["url"]]
                        if label not in existing["engine"]:
                            existing["engine"] += f" · {label}"
                    else:
                        by_url[r["url"]] = r

    return list(by_url.values())


def parse_results(data: dict) -> list[dict]:
    """Extract results from a single SerpAPI response (used by CLI)."""
    return _parse_engine_results(data, "Google Lens")


# ── Exports ───────────────────────────────────────────────────────────────────

def export_csv(results: list[dict], path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["engine", "title", "url", "source", "thumbnail"], extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)
    print(f"  CSV   → {path}")


def _fetch_image_bytes(url: str) -> BytesIO | None:
    try:
        resp = requests.get(url, timeout=8)
        if resp.status_code == 200:
            return BytesIO(resp.content)
    except Exception:
        pass
    return None


def export_excel(results: list[dict], path: str) -> None:
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("  Excel → skipped (run: pip install openpyxl)")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header row
    headers = ["#", "Thumbnail", "Title", "URL", "Source", "Engine"]
    header_fill = PatternFill("solid", fgColor="222222")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 65
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 15

    ROW_HEIGHT = 70  # points ≈ 93 px

    for i, result in enumerate(results, 1):
        row = i + 1
        ws.row_dimensions[row].height = ROW_HEIGHT

        ws.cell(row=row, column=1, value=i)

        title_cell = ws.cell(row=row, column=3, value=result["title"])
        title_cell.alignment = Alignment(wrap_text=True, vertical="center")

        url = result["url"]
        url_cell = ws.cell(row=row, column=4, value=url)
        if url:
            url_cell.hyperlink = url
            url_cell.style = "Hyperlink"
        url_cell.alignment = Alignment(wrap_text=True, vertical="center")

        ws.cell(row=row, column=5, value=result["source"]).alignment = Alignment(vertical="center")
        ws.cell(row=row, column=6, value=result.get("engine", "")).alignment = Alignment(vertical="center")

        thumb_url = result.get("thumbnail", "")
        if thumb_url:
            img_bytes = _fetch_image_bytes(thumb_url)
            if img_bytes:
                try:
                    img = XLImage(img_bytes)
                    img.height = 65
                    img.width = 100
                    img.anchor = f"B{row}"
                    ws.add_image(img)
                except Exception:
                    ws.cell(row=row, column=2, value=thumb_url)
            else:
                ws.cell(row=row, column=2, value=thumb_url)

    wb.save(path)
    print(f"  Excel → {path}")


def export_html(results: list[dict], path: str, source_image_url: str) -> None:
    rows = ""
    for i, r in enumerate(results, 1):
        thumb = r.get("thumbnail", "")
        img_tag = (
            f'<img src="{thumb}" loading="lazy" '
            f'style="max-height:80px;max-width:120px;" '
            f'onerror="this.replaceWith(document.createTextNode(\'—\'))">'
            if thumb else "—"
        )
        url = r.get("url", "")
        link = f'<a href="{url}" target="_blank" rel="noopener">{url}</a>' if url else "—"
        rows += (
            f"<tr>"
            f"<td>{i}</td>"
            f"<td class='thumb'>{img_tag}</td>"
            f"<td>{r.get('title', '')}</td>"
            f"<td class='url'>{link}</td>"
            f"<td>{r.get('source', '')}</td>"
            f"<td>{r.get('engine', '')}</td>"
            f"</tr>\n"
        )

    source_img_tag = (
        f'<img src="{source_image_url}" '
        f'style="max-height:150px;border:1px solid #ccc;" '
        f'onerror="this.style.display=\'none\'">'
        if source_image_url.startswith("http") else ""
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Reverse Image Search Results</title>
<style>
  body {{ font-family: Arial, sans-serif; margin: 24px; color: #222; }}
  h1 {{ font-size: 1.3em; margin-bottom: 4px; }}
  .source {{ margin-bottom: 20px; }}
  .source img {{ display: block; margin-top: 8px; }}
  table {{ border-collapse: collapse; width: 100%; table-layout: fixed; }}
  thead {{ position: sticky; top: 0; z-index: 1; }}
  th {{ background: #222; color: #fff; padding: 8px 10px; text-align: left; font-size: .85em; }}
  td {{ border: 1px solid #ddd; padding: 8px 10px; vertical-align: middle; font-size: .85em; }}
  tr:nth-child(even) {{ background: #f7f7f7; }}
  td.thumb {{ width: 130px; text-align: center; }}
  td.url {{ word-break: break-all; }}
  a {{ color: #1a0dab; }}
  col.n {{ width: 40px; }}
  col.thumb {{ width: 130px; }}
  col.title {{ width: 280px; }}
  col.src {{ width: 160px; }}
</style>
</head>
<body>
<h1>Reverse Image Search Results</h1>
<div class="source">
  <strong>Source image:</strong><br>
  {source_img_tag}
  <a href="{source_image_url}">{source_image_url}</a>
</div>
<p><strong>{len(results)}</strong> results found</p>
<table>
  <colgroup>
    <col class="n"><col class="thumb"><col class="title"><col><col class="src">
  </colgroup>
  <thead>
    <tr><th>#</th><th>Thumbnail</th><th>Title</th><th>URL</th><th>Source</th><th>Engine</th></tr>
  </thead>
  <tbody>
{rows}  </tbody>
</table>
</body>
</html>"""

    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  HTML  → {path}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Reverse image search via SerpAPI.\n"
            "Exports results to CSV, Excel (with embedded thumbnails), and HTML."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "image",
        help="URL of the image to reverse-search, OR a local file path (use with --local)",
    )
    parser.add_argument(
        "--local",
        action="store_true",
        help="Treat 'image' as a local file path (uploads via Google Lens)",
    )
    parser.add_argument(
        "--api-key",
        help="SerpAPI key. Falls back to the SERPAPI_KEY environment variable.",
    )
    parser.add_argument(
        "--output",
        default="results",
        help="Output filename prefix (default: results → results.csv / .xlsx / .html)",
    )
    parser.add_argument("--no-excel", action="store_true", help="Skip Excel output")
    parser.add_argument("--no-html",  action="store_true", help="Skip HTML output")
    parser.add_argument(
        "--save-raw",
        action="store_true",
        help="Save the raw SerpAPI JSON response alongside other outputs",
    )
    args = parser.parse_args()

    api_key = args.api_key or os.environ.get("SERPAPI_KEY")
    if not api_key:
        print(
            "Error: SerpAPI key is required.\n"
            "  Pass it with --api-key YOUR_KEY\n"
            "  or set the SERPAPI_KEY environment variable."
        )
        sys.exit(1)

    print(f"\nSearching for: {args.image}")

    if args.local:
        if not os.path.isfile(args.image):
            print(f"Error: file not found: {args.image}")
            sys.exit(1)
        data = upload_and_search(args.image, api_key)
        source_label = args.image
    else:
        data = search_reverse_image(args.image, api_key)
        source_label = args.image

    if args.save_raw:
        raw_path = f"{args.output}_raw.json"
        with open(raw_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        print(f"  Raw   → {raw_path}")

    results = parse_results(data)

    if not results:
        print("No results found.")
        if not args.save_raw:
            raw_path = f"{args.output}_raw.json"
            with open(raw_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            print(f"Raw API response saved to {raw_path} — check for errors or unexpected keys.")
        sys.exit(0)

    print(f"Found {len(results)} results.\n")

    export_csv(results, f"{args.output}.csv")

    if not args.no_excel:
        export_excel(results, f"{args.output}.xlsx")

    if not args.no_html:
        export_html(results, f"{args.output}.html", source_label)

    print("\nDone.")


if __name__ == "__main__":
    main()
